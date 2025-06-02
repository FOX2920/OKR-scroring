import streamlit as st
import requests
from datetime import datetime, timezone, date, timedelta
from collections import defaultdict
import pandas as pd
import json
import time
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

st.set_page_config(page_title="OKR Scoring System", page_icon="📊", layout="wide")

# Apply custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1E3A8A;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #3B82F6;
        margin-top: 2rem;
    }
    .score-value {
        font-size: 1.2rem;
        font-weight: bold;
    }
    .success-text {
        color: green;
        font-weight: bold;
    }
    .warning-text {
        color: orange;
        font-weight: bold;
    }
    .danger-text {
        color: red;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

# Main title
st.markdown("<h1 class='main-header'>OKR Scoring System</h1>", unsafe_allow_html=True)
st.sidebar.markdown("## OKR Configuration")

# Access tokens for different APIs
GOAL_ACCESS_TOKEN = os.getenv("GOAL_ACCESS_TOKEN")
ACCOUNT_ACCESS_TOKEN = os.getenv("ACCOUNT_ACCESS_TOKEN")
GOOGLE_SHEETS_API_URL = os.getenv("GOOGLE_SHEETS_API_URL")

# Helper function to get current quarter start date
def get_current_quarter_start():
    """Get the first day of the first month of the current quarter"""
    today = date.today()
    current_month = today.month
    
    # Determine which quarter we're in and get the start month
    if current_month <= 3:  # Q1: Jan-Mar
        quarter_start_month = 1
    elif current_month <= 6:  # Q2: Apr-Jun
        quarter_start_month = 4
    elif current_month <= 9:  # Q3: Jul-Sep
        quarter_start_month = 7
    else:  # Q4: Oct-Dec
        quarter_start_month = 10
    
    return date(today.year, quarter_start_month, 1)

# Cache functions to improve performance
@st.cache_data(ttl=3600)
def get_cycle_list(access_token):
    url = "https://goal.base.vn/extapi/v1/cycle/list"
    payload = {'access_token': access_token}
    
    try:
        response = requests.post(url, data=payload)
        response.raise_for_status()
        data = response.json()
        
        # Filter quarterly cycles and convert start time
        quarterly_cycles = [
            {
                'name': cycle['name'], 
                'path': cycle['path'], 
                'start_time': datetime.fromtimestamp(float(cycle['start_time']), tz=timezone.utc),
                'formatted_start_time': datetime.fromtimestamp(float(cycle['start_time']), tz=timezone.utc).strftime('%d/%m/%Y')
            } 
            for cycle in data.get('cycles', []) if cycle.get('metatype') == 'quarterly'
        ]
        
        # Sort cycles by start time in descending order (most recent first)
        quarterly_cycles_sorted = sorted(quarterly_cycles, key=lambda x: x['start_time'], reverse=True)
        
        return quarterly_cycles_sorted
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching cycle list: {e}")
        return []

@st.cache_data(ttl=3600)
def get_account(access_token):
    url = f"https://account.base.vn/extapi/v1/users"
    data = {
        "access_token": access_token
    }

    try:
        response = requests.post(url, data=data)
        response.raise_for_status()  # Check for HTTP errors

        json_response = response.json()

        # Check if response is a list instead of dictionary
        if isinstance(json_response, list) and len(json_response) > 0:
            json_response = json_response[0]

        return json_response  # Return normalized response
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}

@st.cache_data(ttl=3600)
def get_checkins(access_token, path, page, domain="base.vn"):
    url = f"https://goal.{domain}/extapi/v1/cycle/checkins"
    data = {
        "access_token": access_token,
        "path": path,
        "page": page
    }

    try:
        response = requests.post(url, data=data)
        response.raise_for_status()  # Check for HTTP errors

        json_response = response.json()

        # Check if response is a list instead of dictionary
        if isinstance(json_response, list) and len(json_response) > 0:
            json_response = json_response[0]

        return json_response  # Return normalized response
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}

@st.cache_data(ttl=3600)
def get_krs(access_token, path, page, domain="base.vn"):
    url = f"https://goal.base.vn/extapi/v1/cycle/krs"
    data = {
        "access_token": access_token,
        "path": path,
        "page": page
    }

    try:
        response = requests.post(url, data=data)
        response.raise_for_status()  # Check for HTTP errors

        json_response = response.json()

        # Check if response is a list instead of dictionary
        if isinstance(json_response, list) and len(json_response) > 0:
            json_response = json_response[0]

        return json_response  # Return normalized response
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}

@st.cache_data(ttl=3600)
def get_cycle_data(access_token, path):
    url = "https://goal.base.vn/extapi/v1/cycle/get.full"
    payload = {
        'access_token': access_token,
        'path': path
    }
    
    try:
        response = requests.post(url, data=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching cycle data: {e}")
        return {}

# Fetch all data for a specific cycle
def fetch_all_data(cycle_path):
    with st.spinner('Fetching data...'):
        # Fetch account data
        account_response = get_account(ACCOUNT_ACCESS_TOKEN)
        if "error" in account_response:
            st.error(f"Error fetching account data: {account_response['error']}")
            return None, None, None, None
        
        if "users" in account_response:
            account_df = pd.DataFrame(account_response["users"])
        else:
            account_df = pd.DataFrame([account_response])
            
        # Fetch checkins data
        all_checkins = []
        page = 1
        while True:
            response_data = get_checkins(GOAL_ACCESS_TOKEN, cycle_path, page)
            
            if "error" in response_data:
                st.error(f"Error fetching checkins: {response_data['error']}")
                break
                
            checkins = response_data.get("checkins", [])
            
            if not checkins:
                break
                
            all_checkins.extend(checkins)
            page += 1
            
        checkin_df = pd.DataFrame(all_checkins)
        
        # Fetch KRs data
        all_krs = []
        page = 1
        while True:
            response_data = get_krs(GOAL_ACCESS_TOKEN, cycle_path, page)
            
            if "error" in response_data:
                st.error(f"Error fetching KRs: {response_data['error']}")
                break
                
            krs_list = response_data.get("krs", [])
            
            if not krs_list:
                break
                
            all_krs.extend(krs_list)
            page += 1
            
        krs_df = pd.DataFrame(all_krs)
        
        # Fetch cycle data
        cycle_data = get_cycle_data(GOAL_ACCESS_TOKEN, cycle_path)
        
        # Process cycle data
        cycle_df = pd.DataFrame()
        if "targets" in cycle_data:
            targets_list = []
            for target in cycle_data["targets"]:
                if "cached_objs" in target:
                    for obj in target["cached_objs"]:
                        if isinstance(obj, dict):
                            targets_list.append(obj)
            
            if targets_list:
                cycle_df = pd.DataFrame(targets_list)
                
        return account_df, checkin_df, krs_df, cycle_df

# Define User class for OKR tracking
class User:
    def __init__(self, user_id, name, co_OKR=1, checkin=0, dich_chuyen_OKR=0, score=0):
        """Initialize a user with basic attributes."""
        self.user_id = str(user_id)
        self.name = name
        self.co_OKR = co_OKR
        self.checkin = checkin
        self.dich_chuyen_OKR = dich_chuyen_OKR
        self.score = score
        self.OKR = {month: 0 for month in range(1, 13)}  # Create OKR dict for months 1-12

    def update_okr(self, month, value):
        if 1 <= month <= 12:
            self.OKR[month] = value

    def calculate_score(self):
        """Calculate score based on criteria: check-in, OKR and OKR movement."""
        score = 0.5

        # Check-in contributes 1 point
        if self.checkin == 1:
            score += 0.5

        # Having OKR contributes 1 point
        if self.co_OKR == 1:
            score += 1

        # OKR movement score
        movement = self.dich_chuyen_OKR

        if movement < 10:
            score += 0.15
        elif 10 <= movement < 25:
            score += 0.25
        elif 26 <= movement < 30:
            score += 0.5
        elif 31 <= movement < 50:
            score += 0.75
        elif 51 <= movement < 80:
            score += 1.25
        elif 81 <= movement < 99:
            score += 1.5
        elif movement >= 100:
            score += 2.5

        self.score = round(score, 2)  # Round to 2 decimal places

    def __repr__(self):
        return (f"User(id={self.user_id}, name={self.name}, co_OKR={self.co_OKR}, "
                f"checkin={self.checkin}, dich_chuyen_OKR={self.dich_chuyen_OKR}, score={self.score}, "
                f"OKR={self.OKR})")

# Define UserManager class for managing users
class UserManager:
    def __init__(self, account_df, krs_df, checkin_df, cycle_df=None):
        """Initialize UserManager, load data from dataframes."""
        self.account_df = account_df
        self.krs_df = krs_df
        self.checkin_df = checkin_df
        self.cycle_df = cycle_df

        # Create user_id → name mapping from account_df
        self.user_name_map = {}
        if not account_df.empty and 'id' in account_df.columns and 'name' in account_df.columns:
            for _, row in account_df.iterrows():
                self.user_name_map[str(row['id'])] = row.get('name', 'Unknown')

        # Create users list
        self.users = self.create_users()

    def create_users(self):
        """Create User list from KRs data, only for users in account."""
        users = {}
        unique_user_ids = set()

        if not self.krs_df.empty and 'user_id' in self.krs_df.columns:
            for _, kr in self.krs_df.iterrows():
                user_id = str(kr.get("user_id"))
                if user_id and user_id not in unique_user_ids and user_id in self.user_name_map:
                    name = self.user_name_map[user_id]
                    users[user_id] = User(user_id, name)
                    unique_user_ids.add(user_id)

        return users

    def update_checkins(self, start_date=None, end_date=None):
        """Check and update check-in status for each user."""
        for user in self.users.values():
            if self.has_weekly_checkins(user.user_id, start_date, end_date):
                user.checkin = 1
    
    def has_weekly_checkins(self, user_id, start_date=None, end_date=None):
        """Kiểm tra xem user có check-in ít nhất 3 tuần trong khoảng thời gian đã chỉ định không."""
        # Set default date range if not provided
        if start_date is None:
            start_date = get_current_quarter_start()
        if end_date is None:
            end_date = date.today()
            
        # Convert to datetime with timezone for comparison
        start_datetime = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end_datetime = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)
        
        checkins = []
        
        # Thu thập tất cả các lần check-in của user từ checkin_df
        if not self.checkin_df.empty and 'user_id' in self.checkin_df.columns and 'day' in self.checkin_df.columns:
            user_checkins = self.checkin_df[self.checkin_df['user_id'].astype(str) == str(user_id)]
            
            for _, entry in user_checkins.iterrows():
                checkin_date = datetime.fromtimestamp(float(entry.get('day')), tz=timezone.utc)
                checkins.append(checkin_date)
        
        # Lọc ra các lần check-in trong khoảng thời gian đã chỉ định
        checkins_in_range = [dt for dt in checkins if start_datetime <= dt <= end_datetime]
        
        if not checkins_in_range:
            return False  # Không có check-in nào trong khoảng thời gian -> False
        
        # Lưu số tuần có check-in
        weekly_checkins = set(dt.isocalendar()[1] for dt in checkins_in_range)
        
        # Kiểm tra xem user đã check-in ít nhất 3 tuần trong khoảng thời gian chưa
        return len(weekly_checkins) >= 3

    def calculate_scores(self):
        """Calculate score for all users."""
        for user in self.users.values():
            user.calculate_score()

    def get_users(self):
        """Return list of all users."""
        return list(self.users.values())

    def update_okr_movement(self):
        """Update OKR movement for each user."""
        if self.cycle_df is None or self.cycle_df.empty:
            return

        avg_goals = self.calculate_avg_goals()
        
        now = datetime.utcnow()
        current_month = now.month
        current_year = now.year

        # Determine previous month
        if current_month == 1:
            prev_month = 12
            prev_year = current_year - 1
        else:
            prev_month = current_month - 1
            prev_year = current_year

        for user in self.users.values():
            user_id = user.user_id
            
            # Get current OKR value from calculations
            current_okr = avg_goals.get(user_id, 0)
            
            # Get OKR value for previous month from Google Sheets
            prev_okr = self.get_okr_from_sheets(user_id, prev_year, prev_month)
            if prev_okr is None:
                prev_okr = 0
                # Add new data to Google Sheets
                self.insert_okr_to_sheets(user_id, prev_year, prev_month, 0)
            
            # If current month is 1, 4, 7, 10 then keep OKR as is
            if current_month in [1, 4, 7, 10]:
                user.dich_chuyen_OKR = current_okr
            else:
                # Calculate OKR change
                user.dich_chuyen_OKR = round(current_okr - prev_okr, 2)
                
            # Update Google Sheets with current OKR value
            self.update_okr_to_sheets(user_id, current_year, current_month, current_okr)

    def calculate_avg_goals(self):
        """Calculate average OKR for each user using real-time data"""
        goals_data = defaultdict(list)

        if not self.cycle_df.empty and 'type' in self.cycle_df.columns and 'user_id' in self.cycle_df.columns:
            goals_df = self.cycle_df[self.cycle_df['type'] == 'goals']
            for _, goal in goals_df.iterrows():
                user_id = str(goal.get("user_id"))
                current_value = float(goal.get("current_value", 0))
                goals_data[user_id].append(current_value)

        # Calculate average values
        avg_goals = {
            user: sum(values) / len(values) if values else 0
            for user, values in goals_data.items()
        }

        return avg_goals

    def get_okr_from_sheets(self, user_id, year, month):
        """Get OKR value from Google Sheets"""
        params = {
            "action": "get",
            "user_id": user_id,
            "year": year,
            "month": month
        }
        try:
            response = requests.get(GOOGLE_SHEETS_API_URL, params=params)
            data = response.json()
            
            # Check if no data, return None
            okr_value = data.get("okr_value")
            if okr_value is None:
                return None  # Return None to indicate data doesn't exist
            return float(okr_value)
        except Exception as e:
            st.warning(f"Error getting OKR from sheets: {e}")
            return None

    def insert_okr_to_sheets(self, user_id, year, month, okr_value=0):
        """Add OKR data to Google Sheets if it doesn't exist"""
        data = {
            "action": "insert",
            "user_id": user_id,
            "year": year,
            "month": month,
            "okr_value": okr_value
        }
        try:
            response = requests.post(GOOGLE_SHEETS_API_URL, json=data)
            return response.text
        except Exception as e:
            st.warning(f"Error inserting OKR to sheets: {e}")
            return None

    def update_okr_to_sheets(self, user_id, year, month, okr_value):
        """Update OKR data in Google Sheets"""
        # Check if record exists
        existing_value = self.get_okr_from_sheets(user_id, year, month)
        
        if existing_value is None:
            # Insert new record
            data = {
                "action": "insert",
                "user_id": user_id,
                "year": year,
                "month": month,
                "okr_value": okr_value
            }
        else:
            # Update existing record
            data = {
                "action": "update",
                "user_id": user_id,
                "year": year,
                "month": month,
                "okr_value": okr_value
            }
            
        try:
            response = requests.post(GOOGLE_SHEETS_API_URL, json=data)
            return response.text
        except Exception as e:
            st.warning(f"Error updating OKR to sheets: {e}")
            return None

# Function to generate data table
def generate_data_table(users):
    # Create a DataFrame from users
    data = []
    for user in users:
        data.append({
            "Name": user.name,
            "Has OKR": "Yes" if user.co_OKR == 1 else "No",
            "Check-in": "Yes" if user.checkin == 1 else "No",
            "OKR Movement": user.dich_chuyen_OKR,
            "Score": user.score
        })
    
    df = pd.DataFrame(data)
    return df

# Add this function to your file
def export_to_excel(users, filename="output1.xlsx"):
    """
    Xuất dữ liệu OKRs của danh sách users ra file Excel với giao diện được cải tiến.

    Yêu cầu:
      - Mỗi user phải có các thuộc tính: name, co_OKR, checkin, dich_chuyen_OKR, score
    """
    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OKRs"

    # Định nghĩa style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # --- Tiêu đề chính ---
    total_columns = 3 + len(users)  # 3 cột cố định + số user
    last_col_letter = get_column_letter(total_columns)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "ĐÁNH GIÁ OKRs THÁNG"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Header (dòng 2) ---
    fixed_headers = ["TT", "Nội dung", "Tự chấm điểm"]
    user_headers = [user.name for user in users]
    headers = fixed_headers + user_headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        # Đặt độ rộng mặc định cho các cột
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:
            ws.column_dimensions[col_letter].width = 70  # Nội dung dài hơn
        elif col_idx == 1:
            ws.column_dimensions[col_letter].width = 5
        else:
            ws.column_dimensions[col_letter].width = 15

    # --- Các dòng tiêu chí (bắt đầu từ dòng 3) ---
    criteria = [
        [1, "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal (Mục tiêu cá nhân + Đường dẫn)", 1],
        [2, "Có Check-in trên base hàng tuần (Mỗi tuần ít nhất 1 lần check-in)", 0.5],
        [3, "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 0.5],
        [4, "Tổng OKRs dịch chuyển trong tháng (so với tháng trước):", ""],
        ["", "Nhỏ hơn 10%", 0.15],
        ["", "Từ 10 - 25%", 0.25],
        ["", "Từ 26 - 30%", 0.5],
        ["", "Từ 31 - 50%", 0.75],
        ["", "Từ 51 - 80%", 1.25],
        ["", "Từ 81% - 99%", 1.5],
        ["", "100% hoặc có đột phá lớn", 2.5],
        [5, "Tổng cộng OKRs", ""]
    ]
    start_row = 3
    for i, row_data in enumerate(criteria):
        row_idx = start_row + i
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            # Đánh dấu cột loại (nếu giá trị đầu tiên là số thứ tự) với màu nền và in đậm
            if col_idx == 1 and isinstance(value, int):
                cell.fill = category_fill
                cell.font = category_font

    # --- Ghi dữ liệu của từng user ---
    # Các user sẽ được hiển thị từ cột 4 trở đi
    for idx, user in enumerate(users, start=1):
        col_idx = 3 + idx  # cột thứ 1-3 đã dành cho tiêu đề cố định
        col_letter = get_column_letter(col_idx)
        # 1. Đánh giá OKRs cá nhân (dòng 3)
        ws.cell(row=3, column=col_idx, value=1 if user.co_OKR == 1 else 0)
        # 2. Check-in hàng tuần (dòng 4)
        ws.cell(row=4, column=col_idx, value=0.5 if user.checkin == 1 else 0)
        # 3. Check-in với người khác (dòng 5)
        ws.cell(row=5, column=col_idx, value=0.5 )

        # 4. Dịch chuyển OKR:
        # Dòng 6 hiển thị % dịch chuyển, các dòng từ 7 đến 13 hiển thị điểm tương ứng
        movement = user.dich_chuyen_OKR
        ws.cell(row=6, column=col_idx, value=f"{movement}%")

        # Xác định điểm dịch chuyển dựa theo % và dòng ghi điểm:
        if movement < 10:
            score_value = 0.15
            movement_row = 7
        elif movement < 26:
            score_value = 0.25
            movement_row = 8
        elif movement < 31:
            score_value = 0.5
            movement_row = 9
        elif movement < 51:
            score_value = 0.75
            movement_row = 10
        elif movement < 81:
            score_value = 1.25
            movement_row = 11
        elif movement < 100:
            score_value = 1.5
            movement_row = 12
        else:
            score_value = 2.5
            movement_row = 13
        ws.cell(row=movement_row, column=col_idx, value=score_value)

        # 5. Tổng điểm: sử dụng công thức SUM từ dòng 3 đến dòng 13
        formula = user.score
        ws.cell(row=14, column=col_idx, value=formula)

        # Áp dụng border và căn giữa cho các ô dữ liệu của user
        for r in range(3, 15):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # --- Freeze panes để cố định header và tiêu chí ---
    ws.freeze_panes = ws["D3"]

    # --- Tự động điều chỉnh độ rộng cột (nếu cần) ---
    # Vòng lặp qua các cột để tính độ rộng dựa trên nội dung
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Return the workbook object
    return wb


# Function to display user metrics
def display_user_metrics(users):
    # Create metrics
    col1, col2, col3, col4 = st.columns(4)
    
    total_users = len(users)
    users_with_checkins = sum(1 for user in users if user.checkin == 1)
    users_with_okr = sum(1 for user in users if user.co_OKR == 1)
    avg_score = sum(user.score for user in users) / total_users if total_users > 0 else 0
    
    with col1:
        st.metric("Total Users", total_users)
    with col2:
        st.metric("Users with Check-ins", users_with_checkins, f"{users_with_checkins/total_users:.0%}" if total_users > 0 else "0%")
    with col3:
        st.metric("Users with OKR", users_with_okr, f"{users_with_okr/total_users:.0%}" if total_users > 0 else "0%")
    with col4:
        st.metric("Average Score", f"{avg_score:.2f}")

# Main application
def main():
    # Get list of quarterly cycles
    quarterly_cycles = get_cycle_list(GOAL_ACCESS_TOKEN)
    
    if not quarterly_cycles:
        st.error("No quarterly cycles found. Please check your access token.")
        return
    
    # Create cycle selection dropdown
    cycle_options = {f"{cycle['name']} ({cycle['formatted_start_time']})": cycle['path'] for cycle in quarterly_cycles}
    selected_cycle_name = st.sidebar.selectbox(
        "Select Quarter",
        options=list(cycle_options.keys()),
        key="cycle_selector"
    )
    
    if selected_cycle_name:
        selected_cycle_path = cycle_options[selected_cycle_name]
        
        # Add date range selector
        st.sidebar.markdown("## Date Range for Check-ins")
        st.sidebar.markdown("*Filter check-ins within the specified date range*")
        
        # Default dates: from current quarter start to today
        default_start_date = get_current_quarter_start()
        default_end_date = date.today()
        
        start_date = st.sidebar.date_input(
            "Start Date",
            value=default_start_date,
            key="start_date"
        )
        
        end_date = st.sidebar.date_input(
            "End Date",
            value=default_end_date,
            key="end_date"
        )
          # Validate date range
        if isinstance(start_date, date) and isinstance(end_date, date) and start_date > end_date:
            st.sidebar.error("Start date must be before or equal to end date!")
            return
        
        if isinstance(start_date, date) and isinstance(end_date, date):
            st.sidebar.info(f"Analyzing check-ins from {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}")
        
        # Add a Calculate button
        if st.sidebar.button("Calculate Scores", key="calculate_btn"):
            st.session_state.calculate_clicked = True
            
            # Fetch data for selected cycle
            account_df, checkin_df, krs_df, cycle_df = fetch_all_data(selected_cycle_path)
            
            if account_df is not None and checkin_df is not None and krs_df is not None:
                # Initialize UserManager
                manager = UserManager(account_df, krs_df, checkin_df, cycle_df)
                
                # Update check-ins with date range
                manager.update_checkins(start_date, end_date)
                
                # Update OKR movement
                manager.update_okr_movement()
                
                # Calculate scores
                manager.calculate_scores()
                
                # Get users
                users = manager.get_users()
                
                # Store users and date range in session state
                st.session_state.users = users
                st.session_state.date_range = (start_date, end_date)
                
                st.success("Scores calculated successfully!")
            else:
                st.error("Error fetching data. Please try again.")
        
        # Display results if calculation was done
        if hasattr(st.session_state, 'calculate_clicked') and st.session_state.calculate_clicked and hasattr(st.session_state, 'users'):
            st.markdown("<h2 class='sub-header'>OKR Scoring Results</h2>", unsafe_allow_html=True)
              # Display the date range used for calculation
            if hasattr(st.session_state, 'date_range'):
                start_str, end_str = st.session_state.date_range
                if isinstance(start_str, date) and isinstance(end_str, date):
                    st.info(f"📅 Check-ins analyzed from **{start_str.strftime('%d/%m/%Y')}** to **{end_str.strftime('%d/%m/%Y')}**")
            
            # Display metrics
            display_user_metrics(st.session_state.users)
            
            # Display data table
            df = generate_data_table(st.session_state.users)
            
            # Sort by score descending
            df = df.sort_values(by="Score", ascending=False)            # Apply styling
            st.markdown("<h3 class='sub-header'>User Scores</h3>", unsafe_allow_html=True)
            st.dataframe(df, use_container_width=True)
            
            # New Excel download button
            excel_wb = export_to_excel(st.session_state.users)
            
            # Save the workbook to a BytesIO object
            excel_buffer = io.BytesIO()
            excel_wb.save(excel_buffer)
            excel_data = excel_buffer.getvalue()
            
            st.download_button(
                "Download Excel",
                excel_data,
                f"okr_scores_{selected_cycle_name.replace(' ', '_')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download-excel"
            )
if __name__ == "__main__":
    main()
